#!/usr/bin/env python3
"""
KAM Budget Import CSV Preparation - CORRECTED VERSION
Reads Revenue and Gross Profit (£) directly from Excel budget files
Excel structure has 3 sections: Revenue, Gross Profit, and Margin %
We need: Revenue__c from Revenue section, Gross_Margin__c from Gross Profit section
"""

import openpyxl
import pandas as pd
from datetime import datetime

def safe_float(val):
    """Safely convert value to float, handling formulas and strings"""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        if val.startswith('=') or not val.strip():
            return 0.0
        try:
            return float(val)
        except ValueError:
            return 0.0
    return 0.0

def get_user_id(username_map):
    """Map KAM names to Salesforce User IDs"""
    return {
        'Jan Ward': '0051o00000AIeioAAD',
        'Katie Alexander': '005Sj000001u9xhIAA'
    }

def get_account_id_mapping():
    """Map Account Names to Salesforce Account IDs"""
    # Jan Ward's accounts
    jan_ward_accounts = {
        'Anchor Construction Logistics': '001Sj00000By98eIAB',
        'CADDICK CONSTRUCTION LIMITED': '0011o00001YgDLsAAN',
        'Madigan Gill Logistics Limited': '0011o00001ctqlbAAA',
        'Molson Coors': '0012400000RIcW2AAL',
        'ROBERTSON CE LIMITED': '0011o00001Z5Qa6AAF',
        'Story Homes Limited': '0012400000RIcXBAA1',
        'TVS Supply Chain Solutions': '0011o00001jMMhzAAG',
        'Vinci Holdings': '0018e00000MZHLrAAP',
        'Wates Group Limited': '0018e00000MZiXFAA1',
        'Wates Residential Construction Limited': '001Sj00000A7pumIAB'
    }

    # Katie Alexander's accounts
    katie_accounts = {
        'Amey Group PLC': '0012400001H3UtzAAF',
        'Amey Infrastructure Wales Limited': '0018e00000Ma84xAAB',
        'Aureos Group': '001Sj00000HdwLiIAJ',
        'BAM Nuttall': '0012400001BdFBAAA3',
        'CBRE MANAGED SERVICES LIMITED': '0014H00003ayKzlQAE',
        'Citizen': '0014H00002k2COvQAM',
        'Novus Property Solutions': '0012400001168KmAAI',
        'OPENREACH LIMITED': '0014H00002QCOgYQAX',
        'SIGMA RETAIL SOLUTIONS LIMITED': '0011o00001hgIR7AAM',
        'BAM CONSTRUCTION LIMITED': '0012400000RIcTpAAL'  # BAM Construct UK
    }

    # Combine all accounts
    all_accounts = {**jan_ward_accounts, **katie_accounts}
    return all_accounts

def process_kam_budget(excel_file, kam_name, kam_user_id, output_csv, account_mapping):
    """Process a KAM budget Excel file and generate import CSV - CORRECTED VERSION"""

    print(f"\n{'='*60}")
    print(f"Processing: {excel_file}")
    print(f"KAM: {kam_name}")
    print(f"{'='*60}")

    # Load with data_only=True to get calculated values instead of formulas
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb.active

    # Find all three sections: Revenue, Gross Profit, Margin
    revenue_start = None
    gross_profit_start = None
    margin_start = None

    for i in range(1, 50):
        cell_val = ws.cell(i, 1).value
        if cell_val:
            cell_str = str(cell_val).strip().lower()
            if cell_str == 'revenue':
                revenue_start = i + 1  # Data starts next row
            elif cell_str == 'gross profit':
                gross_profit_start = i + 1  # Data starts next row
            elif cell_str == 'margin':
                margin_start = i + 1  # Data starts next row

    if not revenue_start or not gross_profit_start:
        print(f"❌ ERROR: Could not find Revenue or Gross Profit sections in {excel_file}")
        return None

    print(f"Revenue section starts at row {revenue_start}")
    print(f"Gross Profit section starts at row {gross_profit_start}")
    if margin_start:
        print(f"Margin section starts at row {margin_start}")

    # Extract month headers (row 1)
    months = []
    for col in range(2, 14):  # Columns B through M (12 months)
        month_val = ws.cell(1, col).value
        if isinstance(month_val, datetime):
            months.append(month_val.strftime('%Y-%m-01'))
        elif month_val:
            months.append(str(month_val))

    print(f"Found {len(months)} months: {months[0]} to {months[-1] if months else 'N/A'}")

    # Extract revenue and gross profit data
    records = []
    account_count = 0

    # Determine where revenue section ends (before "Total" row or before Gross Profit section)
    revenue_end = gross_profit_start - 2  # Usually stops 2 rows before next section

    # Iterate through account rows in Revenue section ONLY
    for row_idx in range(revenue_start, revenue_end):
        account_name = ws.cell(row_idx, 1).value

        if not account_name or account_name in ['Total', 'MoM Change', 'Gross Profit', '']:
            continue

        account_count += 1

        # Get revenue values for this account (columns 2-13) from REVENUE section
        revenues = []
        for col in range(2, 14):
            val = ws.cell(row_idx, col).value
            revenues.append(safe_float(val))

        # Get gross profit values for this account from GROSS PROFIT section
        # The gross profit row corresponds to the same account at the same offset
        gross_profit_row_idx = gross_profit_start + (row_idx - revenue_start)
        gross_profits_gbp = []
        for col in range(2, 14):
            val = ws.cell(gross_profit_row_idx, col).value
            gross_profits_gbp.append(safe_float(val))

        # Get Account ID from mapping
        account_id = account_mapping.get(account_name)
        if not account_id:
            print(f"⚠️  WARNING: Account '{account_name}' not found in mapping - skipping")
            continue

        # Create records for each month
        for i, month in enumerate(months):
            if i < len(revenues) and i < len(gross_profits_gbp):
                revenue = revenues[i]
                gross_margin_gbp = gross_profits_gbp[i]  # Read directly from Excel in £

                record = {
                    'Account__c': account_id,
                    'Budget_Owner__c': kam_user_id,
                    'Type__c': 'KAM Budget',
                    'Service_Group__c': 'ALL',
                    'Period__c': month,
                    'Revenue__c': round(revenue, 2),
                    'Gross_Margin__c': round(gross_margin_gbp, 2),  # Use value directly from Excel
                    'Actual_Revenue__c': '',
                    'Actual_Margin__c': ''
                }
                records.append(record)

    print(f"Extracted data for {account_count} accounts")
    print(f"Generated {len(records)} records ({account_count} accounts × {len(months)} months)")

    # Convert to DataFrame
    df = pd.DataFrame(records)

    # Save to CSV
    df.to_csv(output_csv, index=False)
    print(f"✅ CSV created: {output_csv}")

    # Show sample records
    print(f"\nSample records (first 3):")
    print(df.head(3).to_string(index=False))

    # Summary statistics
    total_revenue = df['Revenue__c'].sum()
    total_margin = df['Gross_Margin__c'].sum()
    avg_margin_pct = (total_margin / total_revenue * 100) if total_revenue > 0 else 0

    print(f"\n{'='*60}")
    print(f"SUMMARY - {kam_name}")
    print(f"{'='*60}")
    print(f"Total Records: {len(records)}")
    print(f"Accounts: {account_count}")
    print(f"Months: {len(months)}")
    print(f"Date Range: {months[0]} to {months[-1]}")
    print(f"Total Budgeted Revenue: £{total_revenue:,.2f}")
    print(f"Total Budgeted Margin: £{total_margin:,.2f}")
    print(f"Average Margin %: {avg_margin_pct:.1f}%")

    return df

def main():
    print("="*60)
    print("KAM BUDGET IMPORT CSV PREPARATION - CORRECTED VERSION")
    print("="*60)

    user_ids = get_user_id({})
    account_mapping = get_account_id_mapping()

    # Process Jan Ward's budget
    jw_df = process_kam_budget(
        excel_file='/home/john/Projects/Salesforce/JW FY26 Budget.xlsx',
        kam_name='Jan Ward',
        kam_user_id=user_ids['Jan Ward'],
        output_csv='/home/john/Projects/Salesforce/Jan_Ward_Budget_FY26_IMPORT_CORRECTED.csv',
        account_mapping=account_mapping
    )

    # Process Katie Alexander's budget
    ka_df = process_kam_budget(
        excel_file='/home/john/Projects/Salesforce/KA FY26 Budget.xlsx',
        kam_name='Katie Alexander',
        kam_user_id=user_ids['Katie Alexander'],
        output_csv='/home/john/Projects/Salesforce/Katie_Alexander_Budget_FY26_IMPORT_CORRECTED.csv',
        account_mapping=account_mapping
    )

    print("\n" + "="*60)
    print("ALL IMPORT FILES READY!")
    print("="*60)
    print("\nGenerated files:")
    print("  1. Jan_Ward_Budget_FY26_IMPORT_CORRECTED.csv")
    print("  2. Katie_Alexander_Budget_FY26_IMPORT_CORRECTED.csv")
    print("\nNext steps:")
    print("  Import CSV files using sf CLI:")
    print("  sf data import bulk --sobject Management_Information__c \\")
    print("    --file Jan_Ward_Budget_FY26_IMPORT_CORRECTED.csv --target-org OldOrg")
    print()
    print("  sf data import bulk --sobject Management_Information__c \\")
    print("    --file Katie_Alexander_Budget_FY26_IMPORT_CORRECTED.csv --target-org OldOrg")

if __name__ == "__main__":
    main()
